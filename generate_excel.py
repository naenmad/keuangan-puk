import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ===== MASTER TRANSACTION DATA =====
# (Periode YYYY-MM, Tahun, Bulan, No, Nama Transaksi, Kategori, Tipe, Nominal)
transactions = []
monthly_data_seed = [
    # 2023
    ("2023-03", 2023, "Maret", 6738000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Akomodasi aksi DPR RI (tolak perpu cipta kerja)", "Aksi & Advokasi", 1500000)
    ]),
    ("2023-04", 2023, "April", 10138000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Service rutin mobil", "Kendaraan", 1500000)
    ]),
    ("2023-05", 2023, "Mei", 13538000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Akomodasi May Day", "May Day", 2000000),
        ("Pembelian kaos May Day dan spanduk", "Perlengkapan & Atribut", 600000)
    ]),
    ("2023-06", 2023, "Juni", 15838000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Pajak mobil", "Kendaraan", 3900000)
    ]),
    ("2023-07", 2023, "Juli", 16838000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Akomodasi aksi PTUN Bandung", "Aksi & Advokasi", 2000000)
    ]),
    ("2023-08", 2023, "Agustus", 19738000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Biaya daftar futsal DPC CUP", "Olahraga & Kegiatan", 300000),
        ("Akomodasi futsal DPC CUP", "Olahraga & Kegiatan", 2000000),
        ("Akomodasi pengawalan sidang JR di MK", "Aksi & Advokasi", 1000000)
    ]),
    ("2023-09", 2023, "September", 21338000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Akomodasi Aksi KEMNAKER (cabut UU Omnibuslaw)", "Aksi & Advokasi", 1500000)
    ]),
    ("2023-10", 2023, "Oktober", 24738000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Service Mobil PUK", "Kendaraan", 1500000),
        ("Service dan cuci AC PUK", "Aset & Operasional", 500000)
    ]),
    ("2023-11", 2023, "November", 27638000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Ganti dua ban mobil PUK", "Kendaraan", 1900000)
    ]),
    ("2023-12", 2023, "Desember", 30638000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Akomodasi aksi upah 2024 PEMDA Karawang", "Aksi & Advokasi", 1500000),
        ("Aksi KEMNAKER", "Aksi & Advokasi", 2000000)
    ]),
    # 2024
    ("2024-01", 2024, "Januari", 32038000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan Duka cita", "Dana Sosial", 300000),
        ("Aksi DPRD Jabar", "Aksi & Advokasi", 1500000),
        ("Akomodasi Bapor longmarch Bandung-Jakarta", "Aksi & Advokasi", 2000000)
    ]),
    ("2024-02", 2024, "Februari", 33438000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Service mobil PUK", "Kendaraan", 1500000),
        ("Akomodasi Aksi PEMPROV JABAR", "Aksi & Advokasi", 1500000)
    ]),
    ("2024-03", 2024, "Maret", 35338000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Akomodasi aksi DPR RI (cabut omnibuslaw)", "Aksi & Advokasi", 2000000),
        ("Sumbangan BUKBER karyawan PT SAI", "Konsumsi & Kegiatan", 1000000),
        ("Bukber pengurus dan DPC FSP LEM SPSI", "Konsumsi & Kegiatan", 300000),
        ("Bukber pengurus dan manajemen", "Konsumsi & Kegiatan", 2000000)
    ]),
    ("2024-04", 2024, "April", 34938000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Dana sosial suka cita dan duka cita", "Dana Sosial", 300000),
        ("THR DPC dan DPD FSP LEM SPSI", "THR & Bonus", 5000000),
        ("THR pengurus PUK PT SAI", "THR & Bonus", 4000000),
        ("Akomodasi pengantaran jenazah anggota ke Banyuwangi", "Dana Sosial", 2000000)
    ]),
    ("2024-05", 2024, "Mei", 28838000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Akomodasi May Day", "May Day", 3000000),
        ("Beli kaos May Day", "Perlengkapan & Atribut", 720000),
        ("Akomodasi menghadiri MUSNIK PUK TMMIN", "Kegiatan Organisasi", 1000000),
        ("Akomodasi RAKERCAB DPC", "Kegiatan Organisasi", 5000000),
        ("Service mobil", "Kendaraan", 1500000)
    ]),
    ("2024-06", 2024, "Juni", 22518000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Bayar pajak mobil PUK", "Kendaraan", 4000000),
        ("Akomodasi aksi DPRD Karawang", "Aksi & Advokasi", 1000000),
        ("Akomodasi aksi DPRD Bandung", "Aksi & Advokasi", 2000000),
        ("Dana konsolidasi/perjuangan", "Aksi & Advokasi", 5000000)
    ]),
    ("2024-07", 2024, "Juli", 15418000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Pembelian aset laptop Advance", "Aset & Operasional", 5400000),
        ("Akomodasi aksi Disnakertrans JABAR", "Aksi & Advokasi", 2000000)
    ]),
    ("2024-08", 2024, "Agustus", 12918000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Pembuatan kaos anggota 50 pcs", "Perlengkapan & Atribut", 3000000),
        ("Akomodasi RAKORNAS BAPOR", "Kegiatan Organisasi", 2000000)
    ]),
    ("2024-09", 2024, "September", 12818000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000)
    ]),
    ("2024-10", 2024, "Oktober", 17718000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Dana pembangunan gedung DPD FSP LEM SPSI", "Pembangunan Gedung", 2500000)
    ]),
    ("2024-11", 2024, "November", 20118000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Servis Rutin Mobil PUK", "Kendaraan", 1500000),
        ("Ganti ban mobil", "Kendaraan", 2700000)
    ]),
    ("2024-12", 2024, "Desember", 20818000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Ganti ban mobil", "Kendaraan", 900000),
        ("Aksi PEMDA", "Aksi & Advokasi", 1500000),
        ("Aksi PEMPROV", "Aksi & Advokasi", 2000000),
        ("Servis printer PUK", "Aset & Operasional", 303000)
    ]),
    # 2025
    ("2025-01", 2025, "Januari", 22015000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Sumbangan suka cita dan duka", "Dana Sosial", 300000),
        ("Kunjungan kerja DPC FSP LEM SPSI", "Kegiatan Organisasi", 500000)
    ]),
    ("2025-02", 2025, "Februari", 26415000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Akomodasi HUT KSPSI Ke 52 - Transport", "Kegiatan Organisasi", 10000000),
        ("Akomodasi HUT KSPSI Ke 52 - Makan", "Konsumsi & Kegiatan", 1800000),
        ("Akomodasi BAPOR", "Olahraga & Kegiatan", 1000000),
        ("Sumbangan suka cita dan duka", "Dana Sosial", 300000)
    ]),
    ("2025-03", 2025, "Maret", 18515000, 18000000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Sumbangan suka cita dan duka", "Dana Sosial", 300000),
        ("Mutasi mobil operasional PUK", "Kendaraan", 15000000),
        ("Bukber PUK dan Manajemen", "Konsumsi & Kegiatan", 3000000)
    ]),
    ("2025-04", 2025, "April", 5442000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Agenda pertemuan Pleno", "Kegiatan Organisasi", 5000000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Pembuatan kaos anggota 50 pcs", "Perlengkapan & Atribut", 3000000),
        ("Sumbangan suka cita dan duka", "Dana Sosial", 300000)
    ]),
    ("2025-05", 2025, "Mei", 3467000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Sumbangan Duka dan suka Cita", "Dana Sosial", 300000),
        ("Akomodasi May Day", "May Day", 5000000)
    ]),
    ("2025-06", 2025, "Juni", 4492000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Pembuatan Jersey futsal 15 pcs", "Perlengkapan & Atribut", 1500000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Service mobil PUK", "Kendaraan", 1000000)
    ]),
    ("2025-07", 2025, "Juli", 8017000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Pajak Mobil PUK", "Kendaraan", 4000000)
    ]),
    ("2025-08", 2025, "Agustus", 10042000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Konsolidasi Pleno dan mancing", "Olahraga & Kegiatan", 5000000)
    ]),
    ("2025-09", 2025, "September", 11067000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000)
    ]),
    ("2025-10", 2025, "Oktober", 17092000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000)
    ]),
    ("2025-11", 2025, "November", 23117000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000)
    ]),
    ("2025-12", 2025, "Desember", 29142000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Pemberian penghargaan karyawan absensi terbaik", "Penghargaan Karyawan", 3500000),
        ("Akomodasi dan biaya MUSCAB DPC FSP LEM SPSI", "Kegiatan Organisasi", 10850000)
    ]),
    # 2026
    ("2026-01", 2026, "Januari", 20817000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Akomodasi aksi upah 2026 Pemda Karawang", "Aksi & Advokasi", 1000000),
        ("Akomodasi aksi Upah 2026 Bandung 2 kali", "Aksi & Advokasi", 3200000),
        ("Dana konsolidasi gugatan upah 2026", "Aksi & Advokasi", 4000000)
    ]),
    ("2026-02", 2026, "Februari", 18642000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000),
        ("Konsolidasi Pleno dan mancing bersama", "Olahraga & Kegiatan", 5000000),
        ("Sumbangan partisipasi Futsal anggota", "Olahraga & Kegiatan", 600000)
    ]),
    ("2026-03", 2026, "Maret", 19067000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan duka cita", "Dana Sosial", 300000),
        ("Penggantian laptop perusahaan", "Aset & Operasional", 4000000),
        ("Perbaikan kaca mobil PUK", "Kendaraan", 1600000)
    ]),
    ("2026-04", 2026, "April", 19492000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka Cita dan Duka Cita", "Dana Sosial", 300000)
    ]),
    ("2026-05", 2026, "Mei", 25517000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 10125000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Suka cita dan duka cita", "Dana Sosial", 300000),
        ("Akomodasi transport dan makan May Day 2026", "May Day", 15870000)
    ]),
    ("2026-06", 2026, "Juni", 15672000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Setor COS TSCUF", "Setor COS TSCUF", 300000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Dana suka cita dan duka cita", "Dana Sosial", 300000),
        ("Akomodasi pengawalan sidang PTUN Bandung", "Aksi & Advokasi", 1000000),
        ("Pembelian stiker pintu sekretariat PUK", "Perlengkapan & Atribut", 300000)
    ]),
    ("2026-07", 2026, "Juli", 21222000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Setor COS TSCUF", "Setor COS TSCUF", 300000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Dana suka cita dan duka cita", "Dana Sosial", 300000),
        ("Pendaftaran Dialog Ketenagakerjaan", "Kegiatan Organisasi", 3000000),
        ("Akomodasi Dialog Ketenagakerjaan", "Kegiatan Organisasi", 1500000),
        ("Akomodasi RAKOR BAPOR", "Kegiatan Organisasi", 500000)
    ]),
    ("2026-08", 2026, "Agustus", 26672000, 20250000, [
        ("Setor COS dan akomodasi pengurus", "Setor COS & Pengurus", 9000000),
        ("Setor COS TSCUF", "Setor COS TSCUF", 300000),
        ("Angsuran Mobil", "Angsuran Mobil", 3800000),
        ("Dana suka cita dan duka cita", "Dana Sosial", 300000),
        ("Pendaftaran futsal DPC CUP", "Olahraga & Kegiatan", 200000),
        ("Akomodasi futsal DPC CUP", "Olahraga & Kegiatan", 1000000),
        ("Pendaftaran acara mancing DPC", "Olahraga & Kegiatan", 200000),
        ("Akomodasi acara mancing DPC", "Olahraga & Kegiatan", 500000)
    ])
]

# Build flat transaction list
for item in monthly_data_seed:
    code, yr, mth, s_awal, masuk, expenses = item
    # 1. Pemasukan entry
    transactions.append({
        "periode": code,
        "tahun": yr,
        "bulan": mth,
        "keterangan": f"Dana COS Anggota Bulan {mth} {yr}",
        "kategori": "Pemasukan COS",
        "tipe": "Pemasukan",
        "nominal": masuk
    })
    # 2. Expense entries
    for exp_name, exp_cat, exp_val in expenses:
        transactions.append({
            "periode": code,
            "tahun": yr,
            "bulan": mth,
            "keterangan": exp_name,
            "kategori": exp_cat,
            "tipe": "Pengeluaran",
            "nominal": exp_val
        })

# ===== STYLES =====
NAVY = "0F2A44"
STEEL_BLUE = "1B4F72"
ACCENT_BLUE = "2980B9"
SOFT_BLUE = "EBF5FB"
SOFT_GREEN = "E8F8F5"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "C0392B"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_GRAY = "D5D8DC"
DARK_TEXT = "1A1A2E"
SUMMARY_BOX_BG = "EAEDED"

font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_subtitle = Font(name="Calibri", size=10, italic=True, color="D5D8DC")
font_th = Font(name="Calibri", size=10, bold=True, color=WHITE)
font_td = Font(name="Calibri", size=10, color=DARK_TEXT)
font_td_bold = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
font_money_green = Font(name="Calibri", size=10, bold=True, color=ACCENT_GREEN)
font_money_red = Font(name="Calibri", size=10, color=ACCENT_RED)
font_kpi_val = Font(name="Calibri", size=14, bold=True, color="0B5345")
font_kpi_lbl = Font(name="Calibri", size=9, bold=True, color="566573")

fill_navy = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
fill_th = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
fill_alt = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_kpi = PatternFill(start_color=SOFT_GREEN, end_color=SOFT_GREEN, fill_type="solid")
fill_kpi_blue = PatternFill(start_color=SOFT_BLUE, end_color=SOFT_BLUE, fill_type="solid")
fill_total = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=BORDER_GRAY),
    right=Side(style='thin', color=BORDER_GRAY),
    top=Side(style='thin', color=BORDER_GRAY),
    bottom=Side(style='thin', color=BORDER_GRAY)
)
kpi_border = Border(
    left=Side(style='medium', color="16A085"),
    right=Side(style='medium', color="16A085"),
    top=Side(style='medium', color="16A085"),
    bottom=Side(style='medium', color="16A085")
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
fmt_rp = '#,##0'

wb = openpyxl.Workbook()

# ============================================================
# 1. SHEET: DASHBOARD & RINGKASAN (Clean Presentation Layer)
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard & KPI"
ws_dash.sheet_properties.tabColor = NAVY

ws_dash.column_dimensions['A'].width = 4
ws_dash.column_dimensions['B'].width = 24
ws_dash.column_dimensions['C'].width = 24
ws_dash.column_dimensions['D'].width = 24
ws_dash.column_dimensions['E'].width = 24
ws_dash.column_dimensions['F'].width = 4

# Header Title
ws_dash.merge_cells('A1:E2')
ws_dash['A1'] = "EXECUTIVE FINANCIAL DASHBOARD — PUK PT SAI"
ws_dash['A1'].font = font_title; ws_dash['A1'].fill = fill_navy; ws_dash['A1'].alignment = align_center
for c in range(1, 6):
    for r in [1, 2]: ws_dash.cell(row=r, column=c).fill = fill_navy

ws_dash.merge_cells('A3:E3')
ws_dash['A3'] = "Semua Kartu KPI dan Grafik Menggunakan Rumus Otomatis Range Bebas (SUM / SUMIF / MAX)"
ws_dash['A3'].font = font_subtitle; ws_dash['A3'].fill = fill_navy; ws_dash['A3'].alignment = align_center
for c in range(1, 6): ws_dash.cell(row=3, column=c).fill = fill_navy

# Top Summary KPI Cards (Rows 5 to 7)
# Card 1: Total Pemasukan
ws_dash.merge_cells('B5:B5'); ws_dash['B5'] = "TOTAL PEMASUKAN KAS"; ws_dash['B5'].font = font_kpi_lbl; ws_dash['B5'].fill = fill_kpi_blue; ws_dash['B5'].alignment = align_center; ws_dash['B5'].border = thin_border
ws_dash.merge_cells('B6:B7'); ws_dash['B6'] = "=SUM(Data_Bulanan!E5:E1000)"; ws_dash['B6'].font = font_kpi_val; ws_dash['B6'].fill = fill_kpi_blue; ws_dash['B6'].alignment = align_center; ws_dash['B6'].number_format = fmt_rp; ws_dash['B6'].border = thin_border

# Card 2: Total Pengeluaran
ws_dash.merge_cells('C5:C5'); ws_dash['C5'] = "TOTAL PENGELUARAN KAS"; ws_dash['C5'].font = font_kpi_lbl; ws_dash['C5'].fill = fill_kpi_blue; ws_dash['C5'].alignment = align_center; ws_dash['C5'].border = thin_border
ws_dash.merge_cells('C6:C7'); ws_dash['C6'] = "=SUM(Data_Bulanan!F5:F1000)"; ws_dash['C6'].font = font_kpi_val; ws_dash['C6'].fill = fill_kpi_blue; ws_dash['C6'].alignment = align_center; ws_dash['C6'].number_format = fmt_rp; ws_dash['C6'].border = thin_border

# Card 3: Total Surplus Kas
ws_dash.merge_cells('D5:D5'); ws_dash['D5'] = "AKUMULASI SURPLUS (B - C)"; ws_dash['D5'].font = font_kpi_lbl; ws_dash['D5'].fill = fill_kpi_blue; ws_dash['D5'].alignment = align_center; ws_dash['D5'].border = thin_border
ws_dash.merge_cells('D6:D7'); ws_dash['D6'] = "=B6-C6"; ws_dash['D6'].font = font_kpi_val; ws_dash['D6'].fill = fill_kpi_blue; ws_dash['D6'].alignment = align_center; ws_dash['D6'].number_format = fmt_rp; ws_dash['D6'].border = thin_border

# Card 4: Saldo Akhir Terkini
ws_dash.merge_cells('E5:E5'); ws_dash['E5'] = "SALDO KAS TERAKHIR"; ws_dash['E5'].font = font_kpi_lbl; ws_dash['E5'].fill = fill_kpi; ws_dash['E5'].alignment = align_center; ws_dash['E5'].border = kpi_border
ws_dash.merge_cells('E6:E7'); ws_dash['E6'] = "=LOOKUP(2,1/(Data_Bulanan!G5:G1000<>\"\"),Data_Bulanan!G5:G1000)"; ws_dash['E6'].font = Font(name="Calibri", size=14, bold=True, color="1E8449"); ws_dash['E6'].fill = fill_kpi; ws_dash['E6'].alignment = align_center; ws_dash['E6'].number_format = fmt_rp; ws_dash['E6'].border = kpi_border

# Rekap Ringkas per Tahun Table on Dashboard (Row 10)
ws_dash.cell(row=9, column=2, value="RINGKASAN PER TAHUN").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
dash_th = ["Tahun", "Total Pemasukan (Rp)", "Total Pengeluaran (Rp)", "Surplus Tahunan (Rp)"]
for i, h in enumerate(dash_th):
    c = ws_dash.cell(row=10, column=i+2, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

years_list = [2023, 2024, 2025, 2026]
d_r = 11
for yi, yr in enumerate(years_list):
    bg = fill_alt if yi % 2 == 0 else fill_white
    # Tahun
    c = ws_dash.cell(row=d_r, column=2, value=yr)
    c.font = font_td_bold; c.fill = bg; c.alignment = align_center; c.border = thin_border
    # Pemasukan = SUMIFS(Data_Bulanan!E:E, Data_Bulanan!B:B, yr)
    c = ws_dash.cell(row=d_r, column=3, value=f"=SUMIFS(Data_Bulanan!E:E, Data_Bulanan!B:B, B{d_r})")
    c.font = font_money_green; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    # Pengeluaran = SUMIFS(Data_Bulanan!F:F, Data_Bulanan!B:B, yr)
    c = ws_dash.cell(row=d_r, column=4, value=f"=SUMIFS(Data_Bulanan!F:F, Data_Bulanan!B:B, B{d_r})")
    c.font = font_money_red; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    # Surplus = C - D
    c = ws_dash.cell(row=d_r, column=5, value=f"=C{d_r}-D{d_r}")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    d_r += 1

# Total baris ringkas tahunan
c = ws_dash.cell(row=d_r, column=2, value="TOTAL")
c.font = font_th; c.fill = fill_total; c.alignment = align_center; c.border = thin_border
c = ws_dash.cell(row=d_r, column=3, value=f"=SUM(C11:C{d_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
c = ws_dash.cell(row=d_r, column=4, value=f"=SUM(D11:D{d_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
c = ws_dash.cell(row=d_r, column=5, value=f"=C{d_r}-D{d_r}")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Charts on Dashboard
chart_bar = BarChart()
chart_bar.type = "col"
chart_bar.title = "Perbandingan Pemasukan vs Pengeluaran Tahunan"
chart_bar.y_axis.title = "Rupiah"
chart_bar.style = 10
chart_bar.width = 18
chart_bar.height = 13

cats_dash = Reference(ws_dash, min_col=2, min_row=11, max_row=14)
pem_dash = Reference(ws_dash, min_col=3, min_row=10, max_row=14)
pen_dash = Reference(ws_dash, min_col=4, min_row=10, max_row=14)
chart_bar.add_data(pem_dash, titles_from_data=True)
chart_bar.add_data(pen_dash, titles_from_data=True)
chart_bar.set_categories(cats_dash)
chart_bar.series[0].graphicalProperties.solidFill = "27AE60"
chart_bar.series[1].graphicalProperties.solidFill = "E74C3C"
chart_bar.legend.position = 'b'
ws_dash.add_chart(chart_bar, "B17")

# ============================================================
# 2. SHEET: DATA_BULANAN (Pure Continuous Monthly Table)
# ============================================================
ws_month = wb.create_sheet("Data_Bulanan")
ws_month.sheet_properties.tabColor = "27AE60"

ws_month.column_dimensions['A'].width = 12
ws_month.column_dimensions['B'].width = 8
ws_month.column_dimensions['C'].width = 14
ws_month.column_dimensions['D'].width = 20
ws_month.column_dimensions['E'].width = 20
ws_month.column_dimensions['F'].width = 22
ws_month.column_dimensions['G'].width = 20
ws_month.column_dimensions['H'].width = 20
ws_month.column_dimensions['I'].width = 14

ws_month.merge_cells('A1:I2')
ws_month['A1'] = "TABEL KEUANGAN BULANAN (KONTINU & MURNI)"
ws_month['A1'].font = font_title; ws_month['A1'].fill = fill_navy; ws_month['A1'].alignment = align_center
for c in range(1, 10):
    for r in [1, 2]: ws_month.cell(row=r, column=c).fill = fill_navy

ws_month.merge_cells('A3:I3')
ws_month['A3'] = "Tabel bersih tanpa baris subtotal pemutus. Tambah baris baru di bawahnya tanpa merusak rumus!"
ws_month['A3'].font = font_subtitle; ws_month['A3'].fill = fill_navy; ws_month['A3'].alignment = align_center
for c in range(1, 10): ws_month.cell(row=3, column=c).fill = fill_navy

month_headers = ["Periode", "Tahun", "Bulan", "Saldo Awal (Rp)", "Pemasukan (Rp)", "Total Pengeluaran (Rp)", "Saldo Akhir (Rp)", "Surplus/Defisit (Rp)", "Status"]
for i, h in enumerate(month_headers):
    c = ws_month.cell(row=4, column=i+1, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

# Populate continuous rows (starts at row 5)
m_row = 5
for idx, item in enumerate(monthly_data_seed):
    code, yr, mth, s_awal, masuk, exp_list = item
    bg = fill_alt if idx % 2 == 0 else fill_white
    
    # Col A: Periode Code
    c = ws_month.cell(row=m_row, column=1, value=code)
    c.font = font_td_bold; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Col B: Tahun
    c = ws_month.cell(row=m_row, column=2, value=yr)
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Col C: Bulan
    c = ws_month.cell(row=m_row, column=3, value=mth)
    c.font = font_td; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Col D: Saldo Awal
    # Row 1 is initial amount, subsequent rows = G{m_row - 1} (previous month ending balance)
    if idx == 0:
        c = ws_month.cell(row=m_row, column=4, value=s_awal)
    else:
        c = ws_month.cell(row=m_row, column=4, value=f"=G{m_row-1}")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Col E: Pemasukan = SUMIFS from Data_Transaksi
    c = ws_month.cell(row=m_row, column=5, value=f"=SUMIFS(Data_Transaksi!H:H, Data_Transaksi!A:A, A{m_row}, Data_Transaksi!G:G, \"Pemasukan\")")
    c.font = font_money_green; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Col F: Pengeluaran = SUMIFS from Data_Transaksi
    c = ws_month.cell(row=m_row, column=6, value=f"=SUMIFS(Data_Transaksi!H:H, Data_Transaksi!A:A, A{m_row}, Data_Transaksi!G:G, \"Pengeluaran\")")
    c.font = font_money_red; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Col G: Saldo Akhir = D + E - F
    c = ws_month.cell(row=m_row, column=7, value=f"=D{m_row}+E{m_row}-F{m_row}")
    c.font = Font(name="Calibri", size=10, bold=True, color=NAVY); c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Col H: Surplus/Defisit = E - F
    c = ws_month.cell(row=m_row, column=8, value=f"=E{m_row}-F{m_row}")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Col I: Status = IF(H >= 0, "SURPLUS", "DEFISIT")
    c = ws_month.cell(row=m_row, column=9, value=f'=IF(H{m_row}>=0, "SURPLUS", "DEFISIT")')
    c.font = font_td_bold; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    m_row += 1

num_months_continuous = len(monthly_data_seed)

# Add line chart of Saldo Akhir on Data_Bulanan
chart_saldo = LineChart()
chart_saldo.title = "Grafik Pergerakan Saldo Kas (Otomatis Menyesuaikan)"
chart_saldo.y_axis.title = "Rupiah"
chart_saldo.style = 10
chart_saldo.width = 24
chart_saldo.height = 15

cats_month = Reference(ws_month, min_col=1, min_row=5, max_row=m_row-1)
saldo_vals = Reference(ws_month, min_col=7, min_row=4, max_row=m_row-1)
chart_saldo.add_data(saldo_vals, titles_from_data=True)
chart_saldo.set_categories(cats_month)
chart_saldo.series[0].graphicalProperties.line.solidFill = "2980B9"
chart_saldo.series[0].graphicalProperties.line.width = 26000
chart_saldo.legend.position = 'b'
ws_month.add_chart(chart_saldo, "K4")

# ============================================================
# 3. SHEET: DATA_TRANSAKSI (Pure Ledger / Buku Kas Transaksi)
# ============================================================
ws_tx = wb.create_sheet("Data_Transaksi")
ws_tx.sheet_properties.tabColor = "F39C12"

ws_tx.column_dimensions['A'].width = 12
ws_tx.column_dimensions['B'].width = 8
ws_tx.column_dimensions['C'].width = 12
ws_tx.column_dimensions['D'].width = 6
ws_tx.column_dimensions['E'].width = 46
ws_tx.column_dimensions['F'].width = 24
ws_tx.column_dimensions['G'].width = 14
ws_tx.column_dimensions['H'].width = 20

ws_tx.merge_cells('A1:H2')
ws_tx['A1'] = "BUKU TRANSAKSI HARIAN / POS PENGELUARAN & PEMASUKAN"
ws_tx['A1'].font = font_title; ws_tx['A1'].fill = fill_navy; ws_tx['A1'].alignment = align_center
for c in range(1, 9):
    for r in [1, 2]: ws_tx.cell(row=r, column=c).fill = fill_navy

ws_tx.merge_cells('A3:H3')
ws_tx['A3'] = "Input transaksi baru di baris paling bawah. Data bulanan dan dashboard otomatis mendeteksi tanpa copy-paste!"
ws_tx['A3'].font = font_subtitle; ws_tx['A3'].fill = fill_navy; ws_tx['A3'].alignment = align_center
for c in range(1, 9): ws_tx.cell(row=3, column=c).fill = fill_navy

tx_headers = ["Periode", "Tahun", "Bulan", "No", "Keterangan Transaksi", "Kategori", "Tipe", "Nominal (Rp)"]
for i, h in enumerate(tx_headers):
    c = ws_tx.cell(row=4, column=i+1, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

tx_row = 5
for idx, tx in enumerate(transactions):
    bg = fill_alt if idx % 2 == 0 else fill_white
    
    # Periode (YYYY-MM)
    c = ws_tx.cell(row=tx_row, column=1, value=tx["periode"])
    c.font = font_td_bold; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Tahun
    c = ws_tx.cell(row=tx_row, column=2, value=tx["tahun"])
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Bulan
    c = ws_tx.cell(row=tx_row, column=3, value=tx["bulan"])
    c.font = font_td; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # No
    c = ws_tx.cell(row=tx_row, column=4, value=idx+1)
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Keterangan
    c = ws_tx.cell(row=tx_row, column=5, value=tx["keterangan"])
    c.font = font_td; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Kategori
    c = ws_tx.cell(row=tx_row, column=6, value=tx["kategori"])
    c.font = font_td; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Tipe (Pemasukan / Pengeluaran)
    c = ws_tx.cell(row=tx_row, column=7, value=tx["tipe"])
    c.font = font_td_bold; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Nominal
    c = ws_tx.cell(row=tx_row, column=8, value=tx["nominal"])
    c.font = font_money_green if tx["tipe"] == "Pemasukan" else font_money_red
    c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    tx_row += 1

# Total summary for ledger at bottom of Ledger
c = ws_tx.cell(row=tx_row, column=5, value="TOTAL TRANSAKSI TERCATAT")
c.font = font_th; c.fill = fill_total; c.alignment = align_left; c.border = thin_border
ws_tx.cell(row=tx_row, column=6).fill = fill_total; ws_tx.cell(row=tx_row, column=6).border = thin_border
ws_tx.cell(row=tx_row, column=7).fill = fill_total; ws_tx.cell(row=tx_row, column=7).border = thin_border

c = ws_tx.cell(row=tx_row, column=8, value=f"=SUM(H5:H{tx_row-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# ============================================================
# 4. SHEET: REKAP_PEMASUKAN (Income Breakdown Matrix & Monthly Trends)
# ============================================================
ws_inc = wb.create_sheet("Rekap_Pemasukan")
ws_inc.sheet_properties.tabColor = "27AE60"

ws_inc.column_dimensions['A'].width = 4
ws_inc.column_dimensions['B'].width = 34
ws_inc.column_dimensions['C'].width = 18
ws_inc.column_dimensions['D'].width = 18
ws_inc.column_dimensions['E'].width = 18
ws_inc.column_dimensions['F'].width = 18
ws_inc.column_dimensions['G'].width = 22
ws_inc.column_dimensions['H'].width = 20

# Header Banner
ws_inc.merge_cells('A1:H2')
ws_inc['A1'] = "REKAPITULASI PEMASUKAN KAS PER KATEGORI & PER TAHUN"
ws_inc['A1'].font = font_title; ws_inc['A1'].fill = fill_navy; ws_inc['A1'].alignment = align_center
for c in range(1, 9):
    for r in [1, 2]: ws_inc.cell(row=r, column=c).fill = fill_navy

ws_inc.merge_cells('A3:H3')
ws_inc['A3'] = "Dihitung Otomatis dari Data_Transaksi & Data_Bulanan menggunakan Rumus Dinamis SUMIFS"
ws_inc['A3'].font = font_subtitle; ws_inc['A3'].fill = fill_navy; ws_inc['A3'].alignment = align_center
for c in range(1, 9): ws_inc.cell(row=3, column=c).fill = fill_navy

# Top Summary KPI Cards (Rows 5 to 7)
ws_inc.merge_cells('B5:B5'); ws_inc['B5'] = "TOTAL PEMASUKAN KAS"; ws_inc['B5'].font = font_kpi_lbl; ws_inc['B5'].fill = fill_kpi_blue; ws_inc['B5'].alignment = align_center; ws_inc['B5'].border = thin_border
ws_inc.merge_cells('B6:B7'); ws_inc['B6'] = "=SUM(Data_Bulanan!E5:E1000)"; ws_inc['B6'].font = font_kpi_val; ws_inc['B6'].fill = fill_kpi_blue; ws_inc['B6'].alignment = align_center; ws_inc['B6'].number_format = fmt_rp; ws_inc['B6'].border = thin_border

ws_inc.merge_cells('C5:C5'); ws_inc['C5'] = "RATA-RATA PEMASUKAN / BULAN"; ws_inc['C5'].font = font_kpi_lbl; ws_inc['C5'].fill = fill_kpi_blue; ws_inc['C5'].alignment = align_center; ws_inc['C5'].border = thin_border
ws_inc.merge_cells('C6:C7'); ws_inc['C6'] = "=AVERAGEIF(Data_Bulanan!E5:E1000,\">0\")"; ws_inc['C6'].font = font_kpi_val; ws_inc['C6'].fill = fill_kpi_blue; ws_inc['C6'].alignment = align_center; ws_inc['C6'].number_format = fmt_rp; ws_inc['C6'].border = thin_border

ws_inc.merge_cells('D5:E5'); ws_inc['D5'] = "PEMASUKAN TERTINGGI (BULAN)"; ws_inc['D5'].font = font_kpi_lbl; ws_inc['D5'].fill = fill_kpi_blue; ws_inc['D5'].alignment = align_center; ws_inc['D5'].border = thin_border
ws_inc.cell(row=5, column=5).border = thin_border; ws_inc.cell(row=5, column=5).fill = fill_kpi_blue
ws_inc.merge_cells('D6:E7'); ws_inc['D6'] = "=MAX(Data_Bulanan!E5:E1000)"; ws_inc['D6'].font = font_kpi_val; ws_inc['D6'].fill = fill_kpi_blue; ws_inc['D6'].alignment = align_center; ws_inc['D6'].number_format = fmt_rp; ws_inc['D6'].border = thin_border
for r in range(6, 8):
    for col in [4, 5]:
        ws_inc.cell(row=r, column=col).border = thin_border; ws_inc.cell(row=r, column=col).fill = fill_kpi_blue

ws_inc.merge_cells('F5:G5'); ws_inc['F5'] = "TARIF COS BULANAN TERAKHIR"; ws_inc['F5'].font = font_kpi_lbl; ws_inc['F5'].fill = fill_kpi; ws_inc['F5'].alignment = align_center; ws_inc['F5'].border = kpi_border
ws_inc.cell(row=5, column=7).border = kpi_border; ws_inc.cell(row=5, column=7).fill = fill_kpi
ws_inc.merge_cells('F6:G7'); ws_inc['F6'] = "=LOOKUP(2,1/(Data_Bulanan!E5:E1000<>\"\"),Data_Bulanan!E5:E1000)"; ws_inc['F6'].font = Font(name="Calibri", size=14, bold=True, color="1E8449"); ws_inc['F6'].fill = fill_kpi; ws_inc['F6'].alignment = align_center; ws_inc['F6'].number_format = fmt_rp; ws_inc['F6'].border = kpi_border
for r in range(6, 8):
    for col in [6, 7]:
        ws_inc.cell(row=r, column=col).border = kpi_border; ws_inc.cell(row=r, column=col).fill = fill_kpi

# --- SECTION 1: REKAP SUMBER / KATEGORI PEMASUKAN ---
ws_inc.cell(row=9, column=2, value="I. REKAPITULASI BERDASARKAN SUMBER / KATEGORI PEMASUKAN").font = Font(name="Calibri", size=11, bold=True, color=NAVY)

inc_headers = ["No", "Sumber / Kategori Pemasukan", "2023 (Rp)", "2024 (Rp)", "2025 (Rp)", "2026 (Rp)", "Total Keseluruhan (Rp)", "Kontribusi %"]
for i, h in enumerate(inc_headers):
    c = ws_inc.cell(row=10, column=i+1, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

unique_income_categories = [
    "Pemasukan COS",
    "Iuran Sukarela & Sumbangan",
    "Bantuan & Subsidi Organisasi",
    "Hasil Usaha & Merchandise",
    "Pendapatan Lain-lain / Bunga Kas"
]

inc_r = 11
tot_inc_cat_row = 11 + len(unique_income_categories)
cat_years = [2023, 2024, 2025, 2026]

for ci, cat_name in enumerate(unique_income_categories):
    bg = fill_alt if ci % 2 == 0 else fill_white
    
    # No
    c = ws_inc.cell(row=inc_r, column=1, value=ci+1)
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Kategori Name
    c = ws_inc.cell(row=inc_r, column=2, value=cat_name)
    c.font = font_td_bold; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Yearly columns (Col C to F) using SUMIFS
    for yi, yr in enumerate(cat_years):
        c = ws_inc.cell(row=inc_r, column=3 + yi, value=f"=SUMIFS(Data_Transaksi!$H:$H, Data_Transaksi!$F:$F, $B{inc_r}, Data_Transaksi!$B:$B, {yr}, Data_Transaksi!$G:$G, \"Pemasukan\")")
        c.font = font_money_green; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Row Total (Col G) = SUM(C{inc_r}:F{inc_r})
    c = ws_inc.cell(row=inc_r, column=7, value=f"=SUM(C{inc_r}:F{inc_r})")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Kontribusi % (Col H) = G{inc_r} / Grand Total
    c = ws_inc.cell(row=inc_r, column=8, value=f"=IF($G${tot_inc_cat_row}>0, G{inc_r}/$G${tot_inc_cat_row}, 0)")
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border; c.number_format = '0.0%'
    
    inc_r += 1

# Total Category Row
c = ws_inc.cell(row=inc_r, column=2, value="TOTAL PEMASUKAN KAS")
c.font = font_th; c.fill = fill_total; c.alignment = align_left; c.border = thin_border
ws_inc.cell(row=inc_r, column=1).fill = fill_total; ws_inc.cell(row=inc_r, column=1).border = thin_border

for yi in range(4):
    col_letter = chr(ord('C') + yi)
    c = ws_inc.cell(row=inc_r, column=3 + yi, value=f"=SUM({col_letter}11:{col_letter}{inc_r-1})")
    c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Total All Years (Col G)
c = ws_inc.cell(row=inc_r, column=7, value=f"=SUM(G11:G{inc_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Total % (Col H)
c = ws_inc.cell(row=inc_r, column=8, value=f"=SUM(H11:H{inc_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_center; c.border = thin_border; c.number_format = '0%'

# --- SECTION 2: REKAPITULASI PEMASUKAN BULANAN (JAN - DES) ---
m_sec_start = inc_r + 3
ws_inc.cell(row=m_sec_start, column=2, value="II. REKAPITULASI PEMASUKAN BULANAN (JANUARI — DESEMBER)").font = Font(name="Calibri", size=11, bold=True, color=NAVY)

m_tbl_h = m_sec_start + 1
m_inc_headers = ["No", "Bulan Transaksi", "2023 (Rp)", "2024 (Rp)", "2025 (Rp)", "2026 (Rp)", "Total Akumulasi (Rp)", "Rata-rata Bulanan (Rp)"]
for i, h in enumerate(m_inc_headers):
    c = ws_inc.cell(row=m_tbl_h, column=i+1, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

months_all = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

m_row_curr = m_tbl_h + 1
m_start_data_row = m_row_curr
for mi, m_name in enumerate(months_all):
    bg = fill_alt if mi % 2 == 0 else fill_white
    
    # No
    c = ws_inc.cell(row=m_row_curr, column=1, value=mi+1)
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Bulan
    c = ws_inc.cell(row=m_row_curr, column=2, value=m_name)
    c.font = font_td_bold; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Yearly values from Data_Bulanan
    for yi, yr in enumerate(cat_years):
        c = ws_inc.cell(row=m_row_curr, column=3 + yi, value=f"=SUMIFS(Data_Bulanan!$E:$E, Data_Bulanan!$B:$B, {yr}, Data_Bulanan!$C:$C, $B{m_row_curr})")
        c.font = font_money_green; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Total Akumulasi (Col G)
    c = ws_inc.cell(row=m_row_curr, column=7, value=f"=SUM(C{m_row_curr}:F{m_row_curr})")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Rata-rata Bulanan (Col H)
    c = ws_inc.cell(row=m_row_curr, column=8, value=f"=IFERROR(AVERAGEIF(C{m_row_curr}:F{m_row_curr}, \">0\"), 0)")
    c.font = font_td; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    m_row_curr += 1

m_end_data_row = m_row_curr - 1

# Total Monthly Matrix Row
c = ws_inc.cell(row=m_row_curr, column=2, value="TOTAL PEMASUKAN BULANAN")
c.font = font_th; c.fill = fill_total; c.alignment = align_left; c.border = thin_border
ws_inc.cell(row=m_row_curr, column=1).fill = fill_total; ws_inc.cell(row=m_row_curr, column=1).border = thin_border

for yi in range(4):
    col_letter = chr(ord('C') + yi)
    c = ws_inc.cell(row=m_row_curr, column=3 + yi, value=f"=SUM({col_letter}{m_start_data_row}:{col_letter}{m_end_data_row})")
    c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Total All Years Monthly (Col G)
c = ws_inc.cell(row=m_row_curr, column=7, value=f"=SUM(G{m_start_data_row}:G{m_end_data_row})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Average All Years (Col H)
c = ws_inc.cell(row=m_row_curr, column=8, value=f"=IFERROR(AVERAGEIF(H{m_start_data_row}:H{m_end_data_row}, \">0\"), 0)")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# --- CHARTS ON REKAP PEMASUKAN ---
# Chart 1: Bar Chart Yearly Income Comparison
chart_inc_yr = BarChart()
chart_inc_yr.type = "col"
chart_inc_yr.title = "Perbandingan Total Pemasukan Kas Tiap Tahun (2023 — 2026)"
chart_inc_yr.y_axis.title = "Rupiah"
chart_inc_yr.style = 10
chart_inc_yr.width = 18
chart_inc_yr.height = 14

cats_inc_yr = Reference(ws_dash, min_col=2, min_row=11, max_row=14)
data_inc_yr = Reference(ws_dash, min_col=3, min_row=10, max_row=14)
chart_inc_yr.add_data(data_inc_yr, titles_from_data=True)
chart_inc_yr.set_categories(cats_inc_yr)
chart_inc_yr.series[0].graphicalProperties.solidFill = "27AE60"
chart_inc_yr.legend.position = 'b'
ws_inc.add_chart(chart_inc_yr, f"B{m_row_curr + 3}")

# Chart 2: Clustered Bar Trend Pemasukan Bulanan (Jan - Des)
chart_inc_monthly = BarChart()
chart_inc_monthly.type = "col"
chart_inc_monthly.grouping = "clustered"
chart_inc_monthly.title = "Tren Pemasukan Kas per Bulan Antar Tahun (Januari — Desember)"
chart_inc_monthly.y_axis.title = "Rupiah"
chart_inc_monthly.style = 10
chart_inc_monthly.width = 24
chart_inc_monthly.height = 14

cats_inc_m = Reference(ws_inc, min_col=2, min_row=m_start_data_row, max_row=m_end_data_row)
vals_inc_m = Reference(ws_inc, min_col=3, min_row=m_tbl_h, max_row=m_end_data_row, max_col=6)
chart_inc_monthly.add_data(vals_inc_m, titles_from_data=True)
chart_inc_monthly.set_categories(cats_inc_m)
chart_inc_monthly.legend.position = 'b'
ws_inc.add_chart(chart_inc_monthly, f"E{m_row_curr + 3}")

# ============================================================
# 5. SHEET: REKAP_KATEGORI (Category Breakdown Per Year Matrix)
# ============================================================
ws_cat = wb.create_sheet("Rekap_Kategori")
ws_cat.sheet_properties.tabColor = "8E44AD"

ws_cat.column_dimensions['A'].width = 4
ws_cat.column_dimensions['B'].width = 34
ws_cat.column_dimensions['C'].width = 18
ws_cat.column_dimensions['D'].width = 18
ws_cat.column_dimensions['E'].width = 18
ws_cat.column_dimensions['F'].width = 18
ws_cat.column_dimensions['G'].width = 22
ws_cat.column_dimensions['H'].width = 14

ws_cat.merge_cells('A1:H2')
ws_cat['A1'] = "REKAPITULASI PENGELUARAN PER KATEGORI & PER TAHUN"
ws_cat['A1'].font = font_title; ws_cat['A1'].fill = fill_navy; ws_cat['A1'].alignment = align_center
for c in range(1, 9):
    for r in [1, 2]: ws_cat.cell(row=r, column=c).fill = fill_navy

ws_cat.merge_cells('A3:H3')
ws_cat['A3'] = "Dihitung Otomatis dari Data_Transaksi menggunakan Rumus SUMIFS(Nominal, Kategori, Tahun, \"Pengeluaran\")"
ws_cat['A3'].font = font_subtitle; ws_cat['A3'].fill = fill_navy; ws_cat['A3'].alignment = align_center
for c in range(1, 9): ws_cat.cell(row=3, column=c).fill = fill_navy

cat_headers = ["No", "Kategori Pengeluaran", "2023 (Rp)", "2024 (Rp)", "2025 (Rp)", "2026 (Rp)", "Total Keseluruhan (Rp)", "Kontribusi %"]
for i, h in enumerate(cat_headers):
    c = ws_cat.cell(row=4, column=i+1, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

unique_categories = [
    "Setor COS & Pengurus",
    "Angsuran Mobil",
    "Aksi & Advokasi",
    "Kendaraan",
    "Kegiatan Organisasi",
    "May Day",
    "Dana Sosial",
    "THR & Bonus",
    "Olahraga & Kegiatan",
    "Perlengkapan & Atribut",
    "Konsumsi & Kegiatan",
    "Aset & Operasional",
    "Pembangunan Gedung",
    "Setor COS TSCUF",
    "Penghargaan Karyawan"
]

cat_r = 5
tot_cat_row = 5 + len(unique_categories)

for ci, cat_name in enumerate(unique_categories):
    bg = fill_alt if ci % 2 == 0 else fill_white
    
    # No
    c = ws_cat.cell(row=cat_r, column=1, value=ci+1)
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border
    
    # Kategori Name
    c = ws_cat.cell(row=cat_r, column=2, value=cat_name)
    c.font = font_td_bold; c.fill = bg; c.alignment = align_left; c.border = thin_border
    
    # Yearly columns (Col C to F) using SUMIFS
    for yi, yr in enumerate(cat_years):
        c = ws_cat.cell(row=cat_r, column=3 + yi, value=f"=SUMIFS(Data_Transaksi!$H:$H, Data_Transaksi!$F:$F, $B{cat_r}, Data_Transaksi!$B:$B, {yr}, Data_Transaksi!$G:$G, \"Pengeluaran\")")
        c.font = font_td; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Row Total (Col G) = SUM(C{cat_r}:F{cat_r})
    c = ws_cat.cell(row=cat_r, column=7, value=f"=SUM(C{cat_r}:F{cat_r})")
    c.font = font_td_bold; c.fill = bg; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp
    
    # Kontribusi % (Col H) = G{cat_r} / Grand Total
    c = ws_cat.cell(row=cat_r, column=8, value=f"=G{cat_r}/$G${tot_cat_row}")
    c.font = font_td; c.fill = bg; c.alignment = align_center; c.border = thin_border; c.number_format = '0.0%'
    
    cat_r += 1

# Total Row at Bottom of Matrix
c = ws_cat.cell(row=cat_r, column=2, value="TOTAL PENGELUARAN")
c.font = font_th; c.fill = fill_total; c.alignment = align_left; c.border = thin_border
ws_cat.cell(row=cat_r, column=1).fill = fill_total; ws_cat.cell(row=cat_r, column=1).border = thin_border

for yi in range(4):
    col_letter = chr(ord('C') + yi)
    c = ws_cat.cell(row=cat_r, column=3 + yi, value=f"=SUM({col_letter}5:{col_letter}{cat_r-1})")
    c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Total All Years (Col G)
c = ws_cat.cell(row=cat_r, column=7, value=f"=SUM(G5:G{cat_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_right; c.border = thin_border; c.number_format = fmt_rp

# Total % (Col H)
c = ws_cat.cell(row=cat_r, column=8, value=f"=SUM(H5:H{cat_r-1})")
c.font = font_th; c.fill = fill_total; c.alignment = align_center; c.border = thin_border; c.number_format = '0%'

# Chart: Stacked Column Comparison per Category across Years
chart_cat_bar = BarChart()
chart_cat_bar.type = "col"
chart_cat_bar.grouping = "stacked"
chart_cat_bar.overlap = 100
chart_cat_bar.title = "Perbandingan Pengeluaran per Kategori Tiap Tahun (2023 - 2026)"
chart_cat_bar.y_axis.title = "Rupiah"
chart_cat_bar.style = 10
chart_cat_bar.width = 30
chart_cat_bar.height = 16

cats_ref_matrix = Reference(ws_cat, min_col=2, min_row=5, max_row=cat_r-1)
vals_ref_matrix = Reference(ws_cat, min_col=3, min_row=4, max_row=cat_r-1, max_col=6)
chart_cat_bar.add_data(vals_ref_matrix, titles_from_data=True)
chart_cat_bar.set_categories(cats_ref_matrix)
chart_cat_bar.legend.position = 'b'
ws_cat.add_chart(chart_cat_bar, f"A{cat_r + 3}")

# Set Clean Sheet Views
for s in [ws_dash, ws_month, ws_tx, ws_inc, ws_cat]:
    s.sheet_view.showGridLines = False

# Save to output file
new_layout_file = "/Users/mac/Summit/Neraca/Laporan Keuangan PUK.xlsx"
wb.save(new_layout_file)
print(f"✅ FILE DENGAN REKAP PEMASUKAN & REKAP KATEGORI BERHASIL DIBUAT:\n{new_layout_file}")


